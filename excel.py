from openpyxl import load_workbook
import re
class Excel:
    filename = 'senders.xlsx'
    wb = load_workbook(filename)
    def __init__(self):
        pass
    def writeToExcel(self, shared_list):
        wb = self.wb

        ws = wb.active
        tab = ws.tables["Senders"]
        b = 1
        print(f'SHARED LIST: {shared_list}')
        x = list(shared_list.keys())
        for key in x:
        

            # add new row to the table extending the reference
            maxCol, maxRow = re.split('[-:]', tab.ref)
            RowLetter, RowNum, _ = re.split('(\d+)', maxRow.strip())
            ColLetter, ColNum, _ = re.split('(\d+)', maxCol.strip())
            newRowNum = (int(RowNum) + 1)
            maxRow = RowLetter+str(newRowNum)
            tab.ref = maxCol + ':' + maxRow
            # print(f"new Row added: {newRowNum}")

            # add data to new row dynamically
            charlist = []
            for c in self.char_range(ColLetter, RowLetter):
                # print(c)
                charlist.append(c)

            # add value to end of list
            i = 0
            while i < len(charlist):
                print(f'PROVIDER:  {key} COUNT: {shared_list[key]} ' )
                ws[f'{charlist[i]}{newRowNum}'] = key
                ws[f'{charlist[i+1]}{newRowNum}'] = shared_list[key]
                ws[f'{charlist[i+2]}{newRowNum}'] = 'False'
                break
        wb.save('senders.xlsx')
        print("SAVED THAT BULLSHIT IN THE EXCEL SPREADSHEET")
    def openExcel():
        pass
    def readExcel():
        pass
    def deleteEmails():
        pass
    def char_range(self,c1, c2):
        """Generates the characters from `c1` to `c2`, inclusive."""
        for c in range(ord(c1), ord(c2)+1):
            yield chr(c)
