import re
import imaplib
import email
import multiprocessing
import concurrent.futures
from time import perf_counter
from collections import Counter
from typing import final
from openpyxl import load_workbook
class Providers:
    def __init__(self, senderemail, count):
        self.senderemail = senderemail
        self.count = count

    def __repr__(self):
        d = dict()
        d["EMAIL"] = self.senderemail
        d["COUNT"] = self.count
        return d

    def getDict(self):
        d = dict()
        d["EMAIL"] = self.senderemail
        d["COUNT"] = self.count
        return d

    def addOne(self):
        self.count += 1

    def getEmail(self):
        return self.senderemail

    def getCount(self):
        return self.count

#global varibales
username = ""
apppassword = ""
gmail_host = 'imap.gmail.com'
mail = imaplib.IMAP4_SSL(gmail_host)
filename = 'senders.xlsx'
wb = load_workbook(filename)

def char_range(c1, c2):
    """Generates the characters from `c1` to `c2`, inclusive."""
    for c in range(ord(c1), ord(c2)+1):
        yield chr(c)

# will write the list of providers that sent email to an excel spreadsheet

def writeToExcel():
    global shared_list, wb

    ws = wb.active
    tab = ws.tables["Senders"]
    b = 1
    print(f'SHARED LIST: {shared_list}')
    x = list(shared_list.keys())
    for key in x:
       

        # add new row to the table extending the reference
        maxCol, maxRow = re.split('[-:]', tab.ref)
        RowLetter, RowNum, _ = re.split('(\d+)', maxRow.strip())
        ColLetter, ColNum, _ = re.split('(\d+)', maxCol.strip())
        newRowNum = (int(RowNum) + 1)
        maxRow = RowLetter+str(newRowNum)
        tab.ref = maxCol + ':' + maxRow
        # print(f"new Row added: {newRowNum}")

        # add data to new row dynamically
        charlist = []
        for c in char_range(ColLetter, RowLetter):
            # print(c)
            charlist.append(c)

        # add value to end of list
        i = 0
        while i < len(charlist):
            print(f'PROVIDER:  {key} COUNT: {shared_list[key]} ' )
            ws[f'{charlist[i]}{newRowNum}'] = key
            ws[f'{charlist[i+1]}{newRowNum}'] = shared_list[key]
            ws[f'{charlist[i+2]}{newRowNum}'] = 'False'
            break
    wb.save('senders.xlsx')
    print("SAVED THAT BULLSHIT IN THE EXCEL SPREADSHEET")

def readPWD():
    global apppassword
    with open('apppassword.txt') as f:
        content = f.read()
        apppassword = content

def sortEmail(chunk):
    global mail, apppassword, username
    mail.login(username, apppassword)
    mail.select("INBOX")
    partial_list = email_proccesor(chunk)
    mail.close()
    mail.logout()
    return partial_list

def email_proccesor(chunk):
    current_email = 0
    shared_list = []
    for num in chunk:

        _, data = mail.fetch(num, '(RFC822)')
        _, bytes_data = data[0]

        # convert the byte data to message
        email_message = email.message_from_bytes(bytes_data)
        if any(x["EMAIL"] == email_message["from"] for x in shared_list):
            for x in shared_list:
                if x["EMAIL"] == email_message["from"]:
                    current_email += 1
                    current_count = int(x["COUNT"])
                    new_count = 1 + current_count
                    x["COUNT"] = new_count
                    break
        else:
            current_email += 1
            provider = Providers(email_message["from"], 1)
            shared_list.append(provider.getDict())
        print(f'Progress: {current_email} of {len(chunk)} Percent Complete: {round(current_email/len(chunk) * 100)}%', end='\r')
    return shared_list

def readEmail():
    global apppassword, username, mail, shared_list

    # establish connection with gmail

    # select Inbox
    mail.login(username, apppassword)
    mail.select("INBOX")
    _, selected_mails = mail.search(None, 'ALL')
    print(
        f'\nTotal Messages in Inbox: {len(selected_mails[0].split())}\n\nREADING EMAILS NOW PLEASE WAIT!!!\n')

    # divide and conquer by using multiprocessing to speed up the search!!
    chunked_list = []

    # check if total amount of emails can be cleanly divided by 10 with no remainder.
    # if there is a remainder then round down aka "floor" the value and add 1
    chunk_size = len(selected_mails[0].split()) // 10 if len(
        selected_mails[0].split()) % 10 == 0 else len(selected_mails[0].split()) // 10 + 1

    # creating the seperate jobs by dividng it up into 10 seperate jobs
    mail.close()
    mail.logout()
    for i in range(0, len(selected_mails[0].split()), chunk_size):
        chunked_list.append(selected_mails[0].split()[i:i+chunk_size])
    start_time = perf_counter()
    chunked_shared_list = []
    with concurrent.futures.ProcessPoolExecutor() as executor:
        for result in executor.map(sortEmail, chunked_list):
            print(f'length of the process dictionaries: {len(result)}')
            for d in result:
                chunked_shared_list.append(d)
    x = True
    i = 0
    print(f'LENGTH: {len(chunked_shared_list)} ')
    final_dict = Counter()
    for d in chunked_shared_list:
        final_dict[f'{d["EMAIL"]}'] += d["COUNT"]
    
    
    finish = perf_counter()
    shared_list = dict(final_dict)
    print(f'Finished in all 10 proccesses in {round(finish - start_time, 2)} second(s)')
    writeToExcel()

def main():
    readPWD()
    readEmail()

if __name__ == "__main__":
    manager = multiprocessing.Manager()
    shared_list = {}
    current_list = 0
    main()
